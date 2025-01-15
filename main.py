import os
from flask import Flask, render_template, request, redirect, url_for, send_from_directory, send_file
from openpyxl import Workbook, load_workbook
from datetime import date
from fpdf import FPDF
import qrcode
import io
from werkzeug.utils import secure_filename



app = Flask(__name__)
UPLOAD_FOLDER = 'static/uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
arquivo_excel = "manutencoes.xlsx"

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

from openpyxl import Workbook, load_workbook
from datetime import date

from openpyxl import Workbook, load_workbook
from datetime import date

class SistemaManutencao:
    def __init__(self, arquivo='manutencoes.xlsx'):
        self.arquivo = arquivo

    def criar_arquivo_excel(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Equipamentos"
        ws1.append(["ID", "Nome", "QRCode"])

        # Dados da tabela em ordem correta
        dados = [
            ('1', "banheiro B1 masculino", '1'),
            ('2', "banheiro B1 masculino - 2", '2'),
            ('3', "banheiro B2 masculino - Exclusivo gerência", '3'),
            ('4', "banheiro B2 feminino - Exclusivo gerência", '4'),
            ('5', "banheiro B3 feminino", '5'),
            ('6', "banheiro B3 feminino - 2", '6'),
            ('7', "banheiro B4 unissex", '7'),
            ('8', "banheiro B4 feminino", '8'),
            ('9', "vestiario B5 feminino - Armários e area de descanso", '9'),
            ('10', "vestiario B5 feminino - Vasos e Chuveiros", '10'),
            ('11', "vestiario B6 feminino - Armários e area de descanso", '11'),
            ('12', "vestiario B6 feminino - Vasos e Chuveiros", '12'),
            ('13', "vestiario B7 masculino - Armários e area de descanso", '13'),
            ('14', "vestiario B7 masculino - Vasos e Chuveiros", '14'),
            ('15', "banheiro B8 masculino", '15'),
            ('16', "banheiro B8 feminino", '16'),
            ('17', "banheiro B9 masculino - trancado", '17'),
            ('18', "banheiro B9 feminino - trancado", '18'),
            ('19', "banheiro B11 masculino - exclusivo", '19'),
            ('20', "banheiro B11 feminino - virou DML", '20'),
            ('21', "banheiro B12 UNISSEX - exclusivo", '21'),
            ('22', "banheiro B12 feminino", '22'),
            ('23', "banheiro B13 UNISSEX - exclusivo", '23'),
            ('24', "banheiro B13 feminino", '24'),
            ('25', "banheiro B14 masculino", '25'),
            ('26', "banheiro B14 feminino", '26'),
            ('27', "banheiro B15 masculino", '27'),
            ('28', "banheiro B15 feminino", '28'),
            ('29', "vestiario B16 masculino - Vasos e Chuveiros", '29'),
            ('30', "vestiario B16 masculino - Armários e area de descanso", '30'),
            ('31', "vestiario B17 feminino - Chuveiros", '31'),
            ('32', "vestiario B17 feminino - Armários e area de descanso", '32'),
            ('33', "vestiario B18 - parte vestiario fem", '33'),
            ('34', "banheiro B18 feminino - virou DML", '34'),
            ('35', "vestiario B19 masculino - trancado", '35'),
            ('36', "vestiario B19 feminino - trancado", '36'),
            ('37', "Banheiro C4 Masculino", '37'),
            ('38', "Banheiro C4 Feminino", '38'),
            ('39', "Banheiro C5 Masculino", '39'),
            ('40', "Banheiro C5 Feminino", '40'),
            ('41', "Banheiro C6 Masculino", '41'),
            ('42', "Banheiro C6 Feminino", '42'),
            ('43', "Banheiro C7 Masculino", '43'),
            ('44', "Banheiro C7 Feminino", '44'),
            ('45', "Banheiro C8 Masculino", '45'),
            ('46', "Banheiro C8 Feminino", '46')
        ]
        
        for row in dados:
            ws1.append(row)

        ws2 = wb.create_sheet(title="Manutencoes")
        ws2.append(["Equipamento_ID", "Data_Manutencao", "Descricao"])
        ws3 = wb.create_sheet(title="Pecas")
        ws3.append(["Equipamento_ID", "Descricao", "Numero_Serie"])
        wb.save(self.arquivo)

    def cadastrar_equipamento(self, nome_equipamento, qr_code):
        wb = load_workbook(self.arquivo)
        ws = wb["Equipamentos"]
        next_id = ws.max_row
        ws.append([next_id, nome_equipamento, qr_code])
        wb.save(self.arquivo)

    def registrar_manutencao(self, qr_code, descricao):
        wb = load_workbook(self.arquivo)
        ws_equip = wb["Equipamentos"]

        equipamento_id = None
        for row in ws_equip.iter_rows(min_row=2, values_only=True):
            if row[2] == qr_code:
                equipamento_id = row[0]
                break

        if equipamento_id is None:
            return False

        ws_manut = wb["Manutencoes"]
        ws_manut.append([equipamento_id, str(date.today()), descricao])
        wb.save(self.arquivo)
        return True

    def consultar_historico_por_qr(self, qr_code):
        wb = load_workbook(self.arquivo)
        ws_equip = wb["Equipamentos"]
        equipamento_id = None
        nome_equipamento = None
        for row in ws_equip.iter_rows(min_row=2, values_only=True):
            if row[2] == qr_code:
                equipamento_id = row[0]
                nome_equipamento = row[1]
                break

        if equipamento_id is None:
            return None

        ws_manut = wb["Manutencoes"]
        manutencoes = []
        for row in ws_manut.iter_rows(min_row=2, values_only=True):
            if row[0] == equipamento_id:
                manutencoes.append({"data": row[1], "descricao": row[2]})

        return {"nome": nome_equipamento, "manutencoes": manutencoes}

    def adicionar_peca_excel(self, equipamento_id, descricao, numero_serie):
        try:
            wb = load_workbook(self.arquivo)
            ws = wb["Pecas"]
            ws.append([equipamento_id, descricao, numero_serie])
            wb.save(self.arquivo)
        except Exception as e:
            print(f"Erro ao adicionar peça ao Excel: {e}")




sistema = SistemaManutencao()


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/criar_arquivo')
def criar_arquivo():
    sistema.criar_arquivo_excel()
    return "Arquivo Excel criado com sucesso! <br><a href='/'>Voltar</a>"

@app.route('/download_planilha')
def download_planilha():
    caminho_arquivo = os.path.join(app.root_path, arquivo_excel)
    return send_file(caminho_arquivo, as_attachment=True, download_name='manutencoes.xlsx')


@app.route('/cadastrar_equipamento', methods=['GET', 'POST'])
def cadastrar_equipamento():
    if request.method == 'POST':
        nome_equipamento = request.form['nome_equipamento']
        qr_code = request.form['qr_code']
        sistema.cadastrar_equipamento(nome_equipamento, qr_code)
        return redirect(url_for('index'))
    return render_template('cadastrar_equipamento.html')

@app.route('/registrar_manutencao', methods=['GET', 'POST'])
def registrar_manutencao():
    checklist_items = [
        "Maçanetas", "Dobradiças", "Divisórias", "Suporte papel Higienico",
        "Suporte papel toalha", "Torneiras", "Estado lavatorio", "Estado Sifão",
        "Estado Espelho", "Quados de avisos", "Estado dos ralos", "Estado dos Vasos",
        "Válvula de descarga", "Tubo retratil vaso", "Assento sanitário", "Estado Mictório",
        "Valvula Descarga mictório", "Estado luminárias", "Espelho", "Interruptor / Tomadas",
        "Mola de porta"
    ]
    
    # Lista atualizada de equipamentos
    equipamentos = [
        {"nome": "banheiro B1 masculino", "qr_code": "1"},
        {"nome": "banheiro B1 masculino - 2", "qr_code": "2"},
        {"nome": "banheiro B2 masculino - Exclusivo gerência", "qr_code": "3"},
        {"nome": "banheiro B2 feminino - Exclusivo gerência", "qr_code": "4"},
        {"nome": "banheiro B3 feminino", "qr_code": "5"},
        {"nome": "banheiro B3 feminino - 2", "qr_code": "6"},
        {"nome": "banheiro B4 unissex", "qr_code": "7"},
        {"nome": "banheiro B4 feminino", "qr_code": "8"},
        {"nome": "vestiario B5 feminino - Armários e area de descanso", "qr_code": "9"},
        {"nome": "vestiario B5 feminino - Vasos e Chuveiros", "qr_code": "10"},
        {"nome": "vestiario B6 feminino - Armários e area de descanso", "qr_code": "11"},
        {"nome": "vestiario B6 feminino - Vasos e Chuveiros", "qr_code": "12"},
        {"nome": "vestiario B7 masculino - Armários e area de descanso", "qr_code": "13"},
        {"nome": "vestiario B7 masculino - Vasos e Chuveiros", "qr_code": "14"},
        {"nome": "banheiro B8 masculino", "qr_code": "15"},
        {"nome": "banheiro B8 feminino", "qr_code": "16"},
        {"nome": "banheiro B9 masculino - trancado", "qr_code": "17"},
        {"nome": "banheiro B9 feminino - trancado", "qr_code": "18"},
        {"nome": "banheiro B11 masculino - exclusivo", "qr_code": "19"},
        {"nome": "banheiro B11 feminino - virou DML", "qr_code": "20"},
        {"nome": "banheiro B12 UNISSEX - exclusivo", "qr_code": "21"},
        {"nome": "banheiro B12 feminino", "qr_code": "22"},
        {"nome": "banheiro B13 UNISSEX - exclusivo", "qr_code": "23"},
        {"nome": "banheiro B13 feminino", "qr_code": "24"},
        {"nome": "banheiro B14 masculino", "qr_code": "25"},
        {"nome": "banheiro B14 feminino", "qr_code": "26"},
        {"nome": "banheiro B15 masculino", "qr_code": "27"},
        {"nome": "banheiro B15 feminino", "qr_code": "28"},
        {"nome": "vestiario B16 masculino - Vasos e Chuveiros", "qr_code": "29"},
        {"nome": "vestiario B16 masculino - Armários e area de descanso", "qr_code": "30"},
        {"nome": "vestiario B17 feminino - Chuveiros", "qr_code": "31"},
        {"nome": "vestiario B17 feminino - Armários e area de descanso", "qr_code": "32"},
        {"nome": "vestiario B18 - parte vestiario fem", "qr_code": "33"},
        {"nome": "banheiro B18 feminino - virou DML", "qr_code": "34"},
        {"nome": "vestiario B19 masculino - trancado", "qr_code": "35"},
        {"nome": "vestiario B19 feminino - trancado", "qr_code": "36"},
        {"nome": "Banheiro C4 Masculino", "qr_code": "37"},
        {"nome": "Banheiro C4 Feminino", "qr_code": "38"},
        {"nome": "Banheiro C5 Masculino", "qr_code": "39"},
        {"nome": "Banheiro C5 Feminino", "qr_code": "40"},
        {"nome": "Banheiro C6 Masculino", "qr_code": "41"},
        {"nome": "Banheiro C6 Feminino", "qr_code": "42"},
        {"nome": "Banheiro C7 Masculino", "qr_code": "43"},
        {"nome": "Banheiro C7 Feminino", "qr_code": "44"},
        {"nome": "Banheiro C8 Masculino", "qr_code": "45"},
        {"nome": "Banheiro C8 Feminino", "qr_code": "46"}
    ]
    
    if request.method == 'POST':
        qr_code = request.form['qr_code']
        checklist = {key: request.form.get(f'checklist[{key}]') for key in checklist_items}
        descricao = "; ".join(f"{item}: {status}" for item, status in checklist.items() if status)
        sucesso = sistema.registrar_manutencao(qr_code, descricao)
        if sucesso:
            return "Manutenção registrada com sucesso! <br><a href='registrar_manutencao'>Voltar</a>"
        else:
            return "QR Code não encontrado! <br><a href='/registrar_manutencao'>Tente novamente</a>"
    return render_template('registrar_manutencao.html', checklist_items=checklist_items, equipamentos=equipamentos)



@app.route('/consultar_historico', methods=['GET', 'POST'])
def consultar_historico():
    if request.method == 'POST':
        qr_code = request.form['qr_code']
        historico = sistema.consultar_historico_por_qr(qr_code)
        if historico:
            return render_template('historico.html', historico=historico)
        else:
            return "QR Code não encontrado! <br><a href='/consultar_historico'>Tente novamente</a>"
    return render_template('consultar_historico.html')

@app.route('/upload_planilha', methods=['GET', 'POST'])
def upload_planilha():
    if request.method == 'POST':
        if 'file' not in request.files:
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            integrar_planilha(filepath)
            return redirect(url_for('index'))
    return render_template('upload_planilha.html')

def integrar_planilha(filepath):
    # Carregar a planilha existente
    wb_principal = load_workbook(arquivo_excel)
    ws_equipamentos_principal = wb_principal["Equipamentos"]
    ws_manutencoes_principal = wb_principal["Manutencoes"]
    
    # Carregar a nova planilha
    wb_novo = load_workbook(filepath)
    ws_equipamentos_novo = wb_novo["Equipamentos"]
    ws_manutencoes_novo = wb_novo["Manutencoes"]
    
    # Integrar dados da planilha nova na planilha principal
    for row in ws_equipamentos_novo.iter_rows(min_row=2, values_only=True):
        ws_equipamentos_principal.append(row)
        
    for row in ws_manutencoes_novo.iter_rows(min_row=2, values_only=True):
        ws_manutencoes_principal.append(row)
    
    # Salvar a planilha principal com os novos dados integrados
    wb_principal.save(arquivo_excel)

def carregar_planilha():
    return load_workbook(arquivo_excel)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls'}


@app.route('/explorador')
def explorador():
    files = os.listdir(app.config['UPLOAD_FOLDER'])
    return render_template('explorador.html', files=files)

@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return redirect(url_for('explorador'))
    file = request.files['file']
    if file.filename == '':
        return redirect(url_for('explorador'))
    if file:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(filepath)
    return redirect(url_for('explorador'))

@app.route('/delete/<filename>', methods=['POST'])
def delete_file(filename):
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(filepath):
        os.remove(filepath)
    return redirect(url_for('explorador'))

@app.route('/ler_qrcode')
def ler_qrcode():
    return render_template('ler_qrcode.html')

@app.route('/gerar_qrcode_pdf', methods=['POST'])
def gerar_qrcode_pdf():
    data = request.get_json()
    qr_code_text = data.get('qr_code')
    
    if not qr_code_text:
        return "QR Code é necessário", 400

    # Gerar QR Code
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(qr_code_text)
    qr.make(fit=True)

    img = qr.make_image(fill='black', back_color='white')

    # Salvar a imagem em um arquivo temporário
    temp_file = 'temp_qrcode.png'
    img.save(temp_file)

    # Criar PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(40, 10, 'QR Code:', ln=True)
    pdf.image(temp_file, x=10, y=20, w=100)

    # Salvar PDF em um arquivo temporário
    temp_pdf = 'temp_qrcode.pdf'
    pdf.output(temp_pdf)

    # Ler o PDF e enviar como resposta
    with open(temp_pdf, 'rb') as f:
        buffer_pdf = f.read()

    # Remover os arquivos temporários
    os.remove(temp_file)
    os.remove(temp_pdf)

    return send_file(io.BytesIO(buffer_pdf), as_attachment=True, download_name='qrcode.pdf', mimetype='application/pdf')

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
